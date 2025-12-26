"""
Excel generation for Soil Health Card
Formatted Excel with multiple tables
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import Dict
import io


def generate_excel(farmer_data: Dict, soil_data: Dict, recommendation_data: Dict) -> bytes:
    """
    Generate Excel format Soil Health Card
    Returns Excel bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Soil Health Card"
    
    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='2E7D32')
    heading_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "SOIL HEALTH CARD"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Government of India - Department of Agriculture"
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # Farmer Information
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "FARMER INFORMATION"
    cell.font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    farmer_info = [
        ('Name', farmer_data.get('name', '')),
        ('Village', farmer_data.get('village', '')),
        ('District', farmer_data.get('district', '')),
        ('State', farmer_data.get('state', '')),
        ('Field Area', f"{farmer_data.get('field_area', '')} hectares"),
        ('Crop', f"{farmer_data.get('crop_name', '')} ({farmer_data.get('crop_season', '')})")
    ]
    
    for label, value in farmer_info:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # Table 1: Soil Health Summary
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "TABLE 1: SOIL HEALTH SUMMARY"
    cell.font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    # Headers
    headers = ['Parameter', 'Value', 'Unit', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = heading_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1
    
    # Soil data
    soil_rows = [
        ('Nitrogen (N)', soil_data.get('nitrogen', 'N/A'), 'mg/kg', get_status(soil_data.get('nitrogen', 0), 20, 40)),
        ('Phosphorus (P)', soil_data.get('phosphorus', 'N/A'), 'mg/kg', get_status(soil_data.get('phosphorus', 0), 15, 35)),
        ('Potassium (K)', soil_data.get('potassium', 'N/A'), 'mg/kg', get_status(soil_data.get('potassium', 0), 20, 35)),
        ('pH', soil_data.get('ph', 'N/A'), '–', get_ph_status(soil_data.get('ph', 7))),
    ]
    
    if soil_data.get('ec'):
        soil_rows.append(('EC', soil_data.get('ec'), 'dS/m', 'Normal' if soil_data.get('ec', 0) < 2 else 'High'))
    if soil_data.get('organic_carbon'):
        soil_rows.append(('Organic Carbon', soil_data.get('organic_carbon'), '%', 'Low' if soil_data.get('organic_carbon', 0) < 0.5 else 'Good'))
    
    for data_row in soil_rows:
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        row += 1
    
    row += 1
    
    # Table 2: Fertilizer Recommendations
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "TABLE 2: FERTILIZER RECOMMENDATIONS"
    cell.font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    # Parse additional data
    import json
    additional_data = {}
    try:
        additional_data_str = recommendation_data.get('additional_data', '{}')
        if isinstance(additional_data_str, str) and '{' in additional_data_str:
            json_match = additional_data_str[additional_data_str.find('{'):]
            additional_data = json.loads(json_match)
    except:
        pass
    
    # Fertilizer headers
    fert_headers = ['Fertilizer', 'Quantity (kg/ha)', 'Quantity (kg/acre)', 'Govt. Price/50kg', 'Est. Cost']
    for col, header in enumerate(fert_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = heading_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row += 1
    
    # Fertilizer data
    fertilizer_details = additional_data.get('fertilizer_details', [])
    for fert in fertilizer_details:
        qty_ha = fert.get('quantity_kg_per_hectare', 0)
        qty_acre = fert.get('quantity_kg_per_acre', 0)
        price_raw = fert.get('price_per_50kg', 0)
        
        # Handle unverified prices (strings like "Govt. price not available")
        if isinstance(price_raw, (int, float)):
            price = price_raw
            cost = int((qty_ha / 50) * price)
            price_display = f"₹{price}"
            cost_display = f"₹{cost}"
        else:
            # Price is a string (not verified)
            price_display = str(price_raw)
            cost_display = "N/A"
        
        fert_row = [
            fert.get('fertilizer', ''),
            f"{qty_ha:.2f}",
            f"{qty_acre:.2f}",
            price_display,
            cost_display
        ]
        
        for col, value in enumerate(fert_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col > 1 else 'left')
        row += 1
    
    # Total
    total_row = ['TOTAL', '', '', '', f"₹{recommendation_data.get('estimated_cost', 0)}"]
    for col, value in enumerate(total_row, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        cell.alignment = Alignment(horizontal='right' if col > 1 else 'left')
    row += 2
    
    # Important Notes
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "IMPORTANT NOTES"
    cell.font = Font(name='Arial', size=12, bold=True)
    row += 1
    
    notes = [
        "• Soil test validity: One cropping season",
        "• Follow split application schedules as recommended",
        "• Apply during appropriate soil moisture conditions",
        "• Use FYM/compost @ 5-10 tonnes/ha",
        "• Contact Krishi Vigyan Kendra (KVK) for assistance"
    ]
    
    for note in notes:
        ws[f'A{row}'] = note
        row += 1
    
    row += 1
    ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws[f'A{row}'].font = Font(italic=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def get_status(value, low, high):
    if value < low:
        return 'Low'
    elif value > high:
        return 'High'
    return 'Medium'


def get_ph_status(ph):
    if ph < 5.5:
        return 'Acidic'
    elif ph > 7.5:
        return 'Alkaline'
    return 'Neutral'
